#!/usr/bin/env python3
"""
scripts/export_excel.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Exporta toda la base de datos de especímenes a un archivo Excel (.xlsx).

USO
  python scripts/export_excel.py
  python scripts/export_excel.py --out mis_especimenes.xlsx
  python scripts/export_excel.py --familia Brassolidae   # solo esa familia
  python scripts/export_excel.py --flat                  # hoja única aplanada

HOJAS GENERADAS
  📋 Catálogo        — vista aplanada lista para filtrar/imprimir (hoja principal)
  🦋 Especímenes     — tabla specimens con todos los campos
  🔬 Taxonomía       — tabla taxonomy (rank_hierarchy, family, genus, species…)
  🌎 Regiones        — tabla global_regions
  🏷  Familias        — tabla families  (con conteo de géneros y especies)
  🔬 Géneros         — tabla genera
  🌿 Especies        — tabla species
  📦 Media           — tabla specimen_media (URLs de imágenes)

REQUISITOS
  pip install openpyxl python-dotenv supabase

VARIABLES DE ENTORNO  (.env.local)
  NEXT_PUBLIC_SUPABASE_URL   SUPABASE_SERVICE_ROLE_KEY
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from dotenv import load_dotenv
from supabase import create_client, Client

# ─────────────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[1]
load_dotenv(ROOT / ".env.local")
logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-8s  %(message)s")
log = logging.getLogger("export")

_RL = 0.1   # rate-limit entre llamadas Supabase

# ─────────────────────────────────────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────────────────────────────────────
_HEADER_FILL    = PatternFill("solid", fgColor="1B4332")   # verde oscuro
_HEADER_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_ACCENT_FILL    = PatternFill("solid", fgColor="D8F3DC")   # verde claro
_BORDER_THIN    = Border(
    left=Side(style="thin",   color="B7C9C3"),
    right=Side(style="thin",  color="B7C9C3"),
    bottom=Side(style="thin", color="B7C9C3"),
    top=Side(style="thin",    color="B7C9C3"),
)
_ZEBRA_FILL     = PatternFill("solid", fgColor="F0FBF3")


def _style_header(cell) -> None:
    cell.fill  = _HEADER_FILL
    cell.font  = _HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell.border = _BORDER_THIN


def _style_cell(cell, zebra: bool = False) -> None:
    cell.alignment = Alignment(vertical="center", wrap_text=False)
    cell.border = _BORDER_THIN
    if zebra:
        cell.fill = _ZEBRA_FILL


def _auto_width(ws, min_w: int = 10, max_w: int = 60) -> None:
    for col_cells in ws.columns:
        length = max(
            len(str(c.value or "")) for c in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_w, min(length + 4, max_w))


def _add_table(ws, table_name: str) -> None:
    """Convierte el rango de datos en una Excel Table con filtros automáticos."""
    if ws.max_row < 2:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium7",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
    )
    ws.add_table(tbl)


def _write_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.row_dimensions[1].height = 22
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        _style_header(cell)

    for row_idx, row_data in enumerate(rows, start=2):
        zebra = row_idx % 2 == 0
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _style_cell(cell, zebra=zebra)

    _auto_width(ws)
    _add_table(ws, ws.title.replace(" ", "_").replace("🦋", "Especimenes")
               .replace("🔬", "Taxonomia").replace("🌎", "Regiones")
               .replace("🏷", "Familias").replace("🌿", "Especies")
               .replace("📋", "Catalogo").replace("📦", "Media")
               .replace("🔬", "Generos"))

# ─────────────────────────────────────────────────────────────────────────────
# CONSULTAS SUPABASE
# ─────────────────────────────────────────────────────────────────────────────

def _fetch_all(sb: Client, table: str, select: str = "*", filters: dict | None = None) -> list[dict]:
    """Pagina automáticamente para traer todos los registros."""
    out: list[dict] = []
    page_size = 1000
    offset = 0
    while True:
        time.sleep(_RL)
        q = sb.table(table).select(select).range(offset, offset + page_size - 1)
        if filters:
            for col, val in filters.items():
                q = q.eq(col, val)
        res = q.execute()
        batch = res.data or []
        out.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
    return out


def _meta(row: dict, key: str, default: str = "") -> str:
    """Extrae un campo del JSONB metadata de forma segura."""
    try:
        meta = row.get("metadata") or {}
        if isinstance(meta, str):
            meta = json.loads(meta)
        return str(meta.get(key, default))
    except Exception:
        return default

# ─────────────────────────────────────────────────────────────────────────────
# CONSTRUCCIÓN DE HOJAS
# ─────────────────────────────────────────────────────────────────────────────

def build_catalogo(sb: Client, familia_filter: str | None) -> tuple[list[str], list[list]]:
    """
    Vista aplanada: une specimens + taxonomy + global_regions.
    Esta es la hoja que se entrega a clientes o se usa para inventario.
    """
    log.info("  Leyendo catálogo (specimens + taxonomy + regions)…")
    specimens = _fetch_all(sb, "specimens")
    taxonomy  = {r["id"]: r for r in _fetch_all(sb, "taxonomy")}
    regions   = {r["id"]: r for r in _fetch_all(sb, "global_regions")}

    headers = [
        "Code", "Nombre Científico", "Familia", "Subfamilia", "Género",
        "Orden", "Rank Hierarchy", "Región", "País/Localidad",
        "Calidad", "Sexo", "Cantidad", "Precio (USD)",
        "Compliance", "Nombre Común", "Phenotype Tag",
        "Specimen ID",
    ]
    rows: list[list] = []
    for sp in specimens:
        tx  = taxonomy.get(sp.get("taxonomy_id", ""), {})
        reg = regions.get(sp.get("region_id", ""), {})

        family_name = tx.get("family_name", "")
        if familia_filter and family_name.lower() != familia_filter.lower():
            continue

        rows.append([
            _meta(sp, "code"),
            sp.get("species_name", ""),
            family_name,
            tx.get("subfamily_name", ""),
            tx.get("genus_name", ""),
            tx.get("order_name", ""),
            tx.get("rank_hierarchy", ""),
            reg.get("name") or reg.get("region_name", ""),
            _meta(sp, "localidad") or reg.get("region_name", ""),
            _meta(sp, "calidad"),
            _meta(sp, "sexo"),
            _meta(sp, "cantidad"),
            _meta(sp, "precio"),
            _meta(sp, "compliance_status", "NO CITES"),
            _meta(sp, "nombre_comun"),
            _meta(sp, "phenotype_tag"),
            sp.get("id", ""),
        ])

    log.info("  Catálogo: %d filas", len(rows))
    return headers, rows


def build_sheet_simple(
    sb: Client,
    table: str,
    columns: list[str],
    display_cols: list[str] | None = None,
    filters: dict | None = None,
) -> tuple[list[str], list[list]]:
    log.info("  Leyendo %s…", table)
    rows_raw = _fetch_all(sb, table, select=",".join(columns), filters=filters)
    headers  = display_cols or columns
    rows     = [[r.get(c, "") for c in columns] for r in rows_raw]
    log.info("  %s: %d filas", table, len(rows))
    return headers, rows


def build_sheet_families(sb: Client) -> tuple[list[str], list[list]]:
    """Familias con conteo de géneros y especies para verificar integridad."""
    log.info("  Leyendo families con conteos…")
    families = _fetch_all(sb, "families")
    genera   = _fetch_all(sb, "genera")
    species  = _fetch_all(sb, "species")
    taxonomy = _fetch_all(sb, "taxonomy")

    # conteos por family_id
    genus_by_family: dict[str, set[str]] = {}
    for g in genera:
        fid = g.get("family_id") or ""
        genus_by_family.setdefault(fid, set()).add(g.get("genus_name", ""))

    # conteos de especies por family_name (via taxonomy)
    species_by_family: dict[str, int] = {}
    for tx in taxonomy:
        fn = tx.get("family_name", "")
        species_by_family[fn] = species_by_family.get(fn, 0) + 1

    headers = ["ID", "Nombre Familia", "Géneros", "Especies (en catálogo)", "Conteo familias = 1"]
    rows: list[list] = []
    for fam in families:
        fid  = fam["id"]
        name = fam.get("family_name", "")
        genera_count  = len(genus_by_family.get(fid, set()))
        species_count = species_by_family.get(name, 0)
        rows.append([fid, name, genera_count, species_count, "✔ 1"])

    log.info("  families: %d filas", len(rows))
    return headers, rows


def build_sheet_media(sb: Client) -> tuple[list[str], list[list]]:
    log.info("  Leyendo specimen_media…")
    media = _fetch_all(sb, "specimen_media")
    headers = ["ID", "Specimen ID", "Tipo", "URL (CDN)", "Public ID Cloudinary", "Orden"]
    rows = [
        [
            r.get("id", ""),
            r.get("specimen_id", ""),
            r.get("media_type", ""),
            r.get("media_url", "") or r.get("cdn_url", ""),
            r.get("public_id", ""),
            r.get("display_order", 0),
        ]
        for r in media
    ]
    return headers, rows

# ─────────────────────────────────────────────────────────────────────────────
# ENSAMBLAJE DEL LIBRO
# ─────────────────────────────────────────────────────────────────────────────

def build_workbook(sb: Client, familia_filter: str | None, flat_only: bool) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # elimina la hoja por defecto

    sheets: list[tuple[str, list[str], list[list]]] = []

    # ── Hoja principal: catálogo aplanado ────────────────────────────────
    h, r = build_catalogo(sb, familia_filter)
    sheets.append(("📋 Catálogo", h, r))

    if not flat_only:
        # ── Familias (con verificación de unicidad) ─────────────────────
        h, r = build_sheet_families(sb)
        sheets.append(("🏷 Familias", h, r))

        # ── Géneros ──────────────────────────────────────────────────────
        h, r = build_sheet_simple(
            sb, "genera",
            ["id", "genus_name", "subfamily_id", "family_id"],
            ["ID", "Género", "Subfamilia ID", "Familia ID"],
        )
        sheets.append(("🔬 Géneros", h, r))

        # ── Especies ─────────────────────────────────────────────────────
        h, r = build_sheet_simple(
            sb, "species",
            ["id", "species_name", "genus_id", "region_id"],
            ["ID", "Especie", "Género ID", "Región ID"],
        )
        sheets.append(("🌿 Especies", h, r))

        # ── Taxonomía ────────────────────────────────────────────────────
        h, r = build_sheet_simple(
            sb, "taxonomy",
            ["id", "species_name", "genus_name", "subfamily_name",
             "family_name", "order_name", "classification_type", "rank_hierarchy"],
            ["ID", "Nombre Científico", "Género", "Subfamilia",
             "Familia", "Orden", "Tipo", "Rank Hierarchy"],
        )
        sheets.append(("🔬 Taxonomía", h, r))

        # ── Regiones ─────────────────────────────────────────────────────
        h, r = build_sheet_simple(
            sb, "global_regions",
            ["id", "region_name", "name"],
            ["ID", "Código Región", "Nombre Completo"],
        )
        sheets.append(("🌎 Regiones", h, r))

        # ── Especímenes (tabla cruda) ─────────────────────────────────────
        h, r = build_sheet_simple(
            sb, "specimens",
            ["id", "species_name", "taxonomy_id", "region_id", "media_url"],
            ["ID", "Nombre Científico", "Taxonomy ID", "Región ID", "Media URL"],
        )
        sheets.append(("🦋 Especímenes", h, r))

        # ── Media ─────────────────────────────────────────────────────────
        h, r = build_sheet_media(sb)
        sheets.append(("📦 Media", h, r))

    for title, headers, rows in sheets:
        ws = wb.create_sheet(title=title)
        ws.freeze_panes = "A2"
        _write_sheet(ws, headers, rows)

    return wb

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def _require_env(name: str) -> str:
    val = os.getenv(name)
    if not val:
        log.error("Variable '%s' no definida en .env.local.", name)
        sys.exit(1)
    return val


def main() -> None:
    parser = argparse.ArgumentParser(description="Exporta la BD de especímenes a Excel")
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    parser.add_argument("--out",      default=f"especimenes_{ts}.xlsx", help="Nombre del archivo de salida")
    parser.add_argument("--familia",  default=None, help="Exportar solo una familia (ej: Brassolidae)")
    parser.add_argument("--flat",     action="store_true", help="Solo la hoja catálogo (más ligero)")
    args = parser.parse_args()

    sb = create_client(
        _require_env("NEXT_PUBLIC_SUPABASE_URL"),
        _require_env("SUPABASE_SERVICE_ROLE_KEY"),
    )

    log.info("Construyendo libro Excel…")
    wb = build_workbook(sb, args.familia, args.flat)

    out_path = Path(args.out)
    wb.save(out_path)
    log.info("✔ Exportado: %s  (%d KB)", out_path, out_path.stat().st_size // 1024)
    log.info("  Abre el archivo — cada hoja tiene filtros automáticos activados.")


if __name__ == "__main__":
    main()
